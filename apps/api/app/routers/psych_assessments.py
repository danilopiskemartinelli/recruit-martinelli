import csv
import io
import uuid
from datetime import datetime, timezone, timedelta

from fastapi import APIRouter, Depends, status, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.psych_assessment import PsychInvite, PsychResponse
from app.models.application import Application
from app.models.user import User
from app.schemas.psych_assessment import (
    PsychInviteCreate, PsychInviteOut, PsychSubmitPayload, PsychResultsOut,
)
from app.core.rbac import require_recruiter
from app.core.exceptions import NotFoundError
from app.core.psych_questions import (
    BIG_FIVE_QUESTIONS, STRESS_QUESTIONS, DEPRESSION_QUESTIONS,
    MOTIVATORS_GROUPS, DISC_DIMENSIONS, QA_QUESTIONS,
)
from app.core.psych_scoring import calculate_scores

router = APIRouter(tags=["psych-assessments"])


# ── Public endpoints (no auth required) ─────────────────────────────────────

@router.get("/psych-assessments/{token}")
async def get_assessment(token: str, db: AsyncSession = Depends(get_db)):
    """Return assessment metadata and question data for the candidate."""
    result = await db.execute(
        select(PsychInvite)
        .where(PsychInvite.token == token)
        .options(selectinload(PsychInvite.response))
    )
    invite = result.scalar_one_or_none()
    if not invite:
        raise HTTPException(status_code=404, detail="Link de assessment inválido.")
    if invite.expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=410, detail="Este link de assessment expirou.")
    if invite.completed_at:
        raise HTTPException(status_code=409, detail="Este assessment já foi concluído.")

    return {
        "token": token,
        "expires_at": invite.expires_at,
        "tests": {
            "big_five": {
                "title": "Big Five",
                "instruction": "Leia as afirmações e pense o quanto elas descrevem você, enumerando-as em termos de intensidade. Não existem respostas certas ou erradas, o importante é que você responda de forma sincera.",
                "scale": {"min": 1, "max": 7, "labels": {
                    "1": "Não me descreve",
                    "4": "Descreve-me na média",
                    "7": "Descreve-me muito bem",
                }},
                "questions": BIG_FIVE_QUESTIONS,
            },
            "stress": {
                "title": "Nível de Estresse",
                "instruction": "Para cada questão, escolha uma das alternativas de acordo com como você se sentiu no último mês.",
                "scale": {"min": 0, "max": 4, "labels": {
                    "0": "Nunca",
                    "1": "Quase nunca",
                    "2": "Algumas vezes",
                    "3": "Com bastante frequência",
                    "4": "Quase sempre",
                }},
                "questions": STRESS_QUESTIONS,
            },
            "depression": {
                "title": "Sintomático de Depressão",
                "instruction": "Assinale as respostas que chegam mais próximo de como você tem se sentido nos últimos 7 dias, não apenas como você tem se sentido hoje.",
                "questions": DEPRESSION_QUESTIONS,
            },
            "motivators": {
                "title": "Motivadores",
                "instruction": "Marque as opções de acordo com o que você realmente sente.",
                "scale": {"min": 0, "max": 5, "labels": {
                    "0": "Nunca",
                    "1": "Quase nunca",
                    "2": "Poucas vezes",
                    "3": "Normalmente",
                    "4": "Quase sempre",
                    "5": "O tempo todo",
                }},
                "groups": MOTIVATORS_GROUPS,
            },
            "disc": {
                "title": "DISC",
                "instruction": "Marque de acordo com o que você pensa sobre você mesmo(a). Marque apenas uma opção por grupo.",
                "dimensions": DISC_DIMENSIONS,
            },
            "qa": {
                "title": "Quociente de Adversidade",
                "instruction": "Marque as opções de acordo com o que você realmente sente.",
                "scale": {"min": 1, "max": 5, "labels": {
                    "1": "Nunca",
                    "2": "Pouco",
                    "3": "Médio",
                    "4": "Na maioria das vezes",
                    "5": "Sempre",
                }},
                "questions": QA_QUESTIONS,
            },
        },
    }


@router.post("/psych-assessments/{token}/submit", status_code=status.HTTP_201_CREATED)
async def submit_assessment(
    token: str,
    payload: PsychSubmitPayload,
    db: AsyncSession = Depends(get_db),
):
    """Receive candidate answers, compute scores, and persist the response."""
    result = await db.execute(
        select(PsychInvite).where(PsychInvite.token == token)
    )
    invite = result.scalar_one_or_none()
    if not invite:
        raise HTTPException(status_code=404, detail="Link de assessment inválido.")
    if invite.expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=410, detail="Este link de assessment expirou.")
    if invite.completed_at:
        raise HTTPException(status_code=409, detail="Este assessment já foi concluído.")

    now = datetime.now(timezone.utc)
    response = PsychResponse(
        invite_id=invite.id,
        big_five_answers=payload.big_five_answers,
        stress_answers=payload.stress_answers,
        depression_answers=payload.depression_answers,
        motivators_answers=payload.motivators_answers,
        disc_answers=payload.disc_answers,
        qa_answers=payload.qa_answers,
        completed_at=now,
    )
    response.scores = calculate_scores(response)
    db.add(response)

    invite.completed_at = now
    await db.flush()

    return {"message": "Assessment concluído com sucesso!", "response_id": response.id}


# ── Recruiter endpoints (auth required) ─────────────────────────────────────

@router.post(
    "/applications/{app_id}/psych-invite",
    response_model=PsychInviteOut,
    status_code=status.HTTP_201_CREATED,
)
async def create_psych_invite(
    app_id: uuid.UUID,
    payload: PsychInviteCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter),
):
    """Create (or return existing) a psych assessment invite for an application."""
    app_result = await db.execute(
        select(Application).where(
            Application.id == str(app_id),
            Application.company_id == current_user.company_id,
        )
    )
    application = app_result.scalar_one_or_none()
    if not application:
        raise NotFoundError("Application not found")

    # Check if there's already an incomplete invite
    existing = await db.execute(
        select(PsychInvite).where(
            PsychInvite.application_id == str(app_id),
            PsychInvite.completed_at.is_(None),
        )
    )
    invite = existing.scalar_one_or_none()
    if not invite:
        invite = PsychInvite(
            application_id=str(app_id),
            expires_at=datetime.now(timezone.utc) + timedelta(days=payload.expires_in_days),
            sent_at=datetime.now(timezone.utc),
        )
        db.add(invite)
        await db.flush()

    base_url = str(request.base_url).rstrip("/")
    assessment_url = f"{base_url}/portal/assessment/{invite.token}"

    out = PsychInviteOut.model_validate(invite)
    out.assessment_url = assessment_url
    return out


@router.get("/applications/{app_id}/psych-results", response_model=PsychResultsOut)
async def get_psych_results(
    app_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter),
):
    """Get the completed psych assessment results for an application."""
    app_result = await db.execute(
        select(Application).where(
            Application.id == str(app_id),
            Application.company_id == current_user.company_id,
        )
    )
    if not app_result.scalar_one_or_none():
        raise NotFoundError("Application not found")

    result = await db.execute(
        select(PsychResponse)
        .join(PsychInvite, PsychResponse.invite_id == PsychInvite.id)
        .where(PsychInvite.application_id == str(app_id))
        .order_by(PsychResponse.completed_at.desc())
    )
    response = result.scalar_one_or_none()
    if not response:
        raise HTTPException(status_code=404, detail="No completed assessment found for this application.")
    return response


@router.get("/applications/{app_id}/psych-answers")
async def get_psych_answers(
    app_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter),
):
    """Return raw answers alongside question text so recruiters can review."""
    app_result = await db.execute(
        select(Application).where(
            Application.id == str(app_id),
            Application.company_id == current_user.company_id,
        )
    )
    if not app_result.scalar_one_or_none():
        raise NotFoundError("Application not found")

    result = await db.execute(
        select(PsychResponse)
        .join(PsychInvite, PsychResponse.invite_id == PsychInvite.id)
        .where(PsychInvite.application_id == str(app_id))
        .order_by(PsychResponse.completed_at.desc())
    )
    response = result.scalar_one_or_none()
    if not response:
        raise HTTPException(status_code=404, detail="No completed assessment found for this application.")

    disc_dims_out = {
        dim: {"name": data["name"], "pairs": [list(p) for p in data["pairs"]]}
        for dim, data in DISC_DIMENSIONS.items()
    }

    return {
        "completed_at": response.completed_at,
        "tests": {
            "big_five": {
                "title": "Big Five",
                "scale": {"min": 1, "max": 7},
                "questions": BIG_FIVE_QUESTIONS,
                "answers": response.big_five_answers,
            },
            "stress": {
                "title": "Nível de Estresse",
                "scale": {"min": 0, "max": 4, "labels": {
                    "0": "Nunca", "1": "Quase nunca", "2": "Algumas vezes",
                    "3": "Com bastante frequência", "4": "Quase sempre",
                }},
                "questions": STRESS_QUESTIONS,
                "answers": response.stress_answers,
            },
            "depression": {
                "title": "Sintomático de Depressão",
                "questions": DEPRESSION_QUESTIONS,
                "answers": response.depression_answers,
            },
            "motivators": {
                "title": "Motivadores",
                "scale": {"min": 0, "max": 5, "labels": {
                    "0": "Nunca", "1": "Quase nunca", "2": "Poucas vezes",
                    "3": "Normalmente", "4": "Quase sempre", "5": "O tempo todo",
                }},
                "groups": MOTIVATORS_GROUPS,
                "answers": response.motivators_answers,
            },
            "disc": {
                "title": "DISC",
                "scale": {"min": 0, "max": 5},
                "dimensions": disc_dims_out,
                "answers": response.disc_answers,
            },
            "qa": {
                "title": "Quociente de Adversidade",
                "scale": {"min": 1, "max": 5, "labels": {
                    "1": "Nunca", "2": "Pouco", "3": "Médio",
                    "4": "Na maioria das vezes", "5": "Sempre",
                }},
                "questions": QA_QUESTIONS,
                "answers": response.qa_answers,
            },
        },
    }


@router.get("/applications/{app_id}/psych-results.csv")
async def export_psych_csv(
    app_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter),
):
    """Export the candidate's raw answers + scores as CSV."""
    from app.models.candidate import Candidate
    app_result = await db.execute(
        select(Application).where(
            Application.id == str(app_id),
            Application.company_id == current_user.company_id,
        )
    )
    application = app_result.scalar_one_or_none()
    if not application:
        raise NotFoundError("Application not found")

    result = await db.execute(
        select(PsychResponse)
        .join(PsychInvite, PsychResponse.invite_id == PsychInvite.id)
        .where(PsychInvite.application_id == str(app_id))
        .order_by(PsychResponse.completed_at.desc())
    )
    response = result.scalar_one_or_none()
    if not response:
        raise HTTPException(status_code=404, detail="No completed assessment found for this application.")

    cand_result = await db.execute(select(Candidate).where(Candidate.id == application.candidate_id))
    candidate = cand_result.scalar_one_or_none()

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["section", "test", "key", "question", "answer", "extra"])

    cand_name = candidate.full_name if candidate else ""
    cand_email = candidate.email if candidate else ""
    w.writerow(["meta", "", "candidate_name", "", cand_name, ""])
    w.writerow(["meta", "", "candidate_email", "", cand_email, ""])
    w.writerow(["meta", "", "completed_at", "", response.completed_at.isoformat() if response.completed_at else "", ""])

    # Big Five
    for q in BIG_FIVE_QUESTIONS:
        ans = response.big_five_answers.get(str(q["num"]), "")
        w.writerow(["answers", "big_five", q["num"], q["text"], ans, f"{q['factor']} / {q['sub_factor']}"])
    # Stress
    for q in STRESS_QUESTIONS:
        ans = response.stress_answers.get(str(q["num"]), "")
        w.writerow(["answers", "stress", q["num"], q["text"], ans, ""])
    # Depression
    for q in DEPRESSION_QUESTIONS:
        ans = response.depression_answers.get(q["key"], "")
        w.writerow(["answers", "depression", q["key"], q["text"], ans, ""])
    # Motivators
    for group, questions in MOTIVATORS_GROUPS.items():
        for q in questions:
            ans = response.motivators_answers.get(str(q["num"]), "")
            w.writerow(["answers", "motivators", q["num"], q["text"], ans, group])
    # DISC
    for dim, info in DISC_DIMENSIONS.items():
        values = response.disc_answers.get(dim, [])
        for idx, pair in enumerate(info["pairs"]):
            v = values[idx] if idx < len(values) else ""
            w.writerow(["answers", "disc", f"{dim}_{idx + 1}", f"{pair[0]} ↔ {pair[1]}", v, f"{dim} - {info['name']}"])
    # QA
    for q in QA_QUESTIONS:
        ans = response.qa_answers.get(str(q["num"]), "")
        w.writerow(["answers", "qa", q["num"], q["text"], ans, ""])

    # Scores
    scores = response.scores or {}
    if "big_five" in scores:
        for factor, data in scores["big_five"].items():
            w.writerow(["scores", "big_five", factor, "", data.get("score", ""), f"{data.get('raw','')}/{data.get('max','')}"])
            for sub, s in data.get("sub_factors", {}).items():
                w.writerow(["scores", "big_five", f"{factor} / {sub}", "", s.get("score", ""), f"{s.get('raw','')}/{s.get('max','')}"])
    if "stress" in scores:
        w.writerow(["scores", "stress", "total", "", scores["stress"].get("score", ""), scores["stress"].get("level", "")])
    if "depression" in scores:
        w.writerow(["scores", "depression", "total", "", scores["depression"].get("score", ""), scores["depression"].get("level", "")])
    for group, m in scores.get("motivators", {}).items():
        w.writerow(["scores", "motivators", group, "", m.get("score", ""), f"{m.get('pct','')}"])
    for dim, d in scores.get("disc", {}).items():
        w.writerow(["scores", "disc", dim, d.get("name", ""), d.get("score", ""), d.get("level", "")])
    if "qa" in scores:
        w.writerow(["scores", "qa", "total", "", scores["qa"].get("score", ""), scores["qa"].get("level", "")])

    buf.seek(0)
    filename = f"psych-results-{cand_name.replace(' ', '_') or app_id}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/applications/{app_id}/psych-results.xlsx")
async def export_psych_xlsx(
    app_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_recruiter),
):
    """Export raw answers + scores as XLSX, one sheet per test for easy validation."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.models.candidate import Candidate

    app_result = await db.execute(
        select(Application).where(
            Application.id == str(app_id),
            Application.company_id == current_user.company_id,
        )
    )
    application = app_result.scalar_one_or_none()
    if not application:
        raise NotFoundError("Application not found")

    result = await db.execute(
        select(PsychResponse)
        .join(PsychInvite, PsychResponse.invite_id == PsychInvite.id)
        .where(PsychInvite.application_id == str(app_id))
        .order_by(PsychResponse.completed_at.desc())
    )
    response = result.scalar_one_or_none()
    if not response:
        raise HTTPException(status_code=404, detail="No completed assessment found for this application.")

    cand_result = await db.execute(select(Candidate).where(Candidate.id == application.candidate_id))
    candidate = cand_result.scalar_one_or_none()

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1c62cb")
    title_font = Font(bold=True, size=14)
    sub_font = Font(italic=True, color="555555")
    wrap = Alignment(wrap_text=True, vertical="top")

    def write_headers(ws, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=ws.max_row, column=i, value=h)
            c.font = header_font
            c.fill = header_fill

    def autosize(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    # ── Resumo ───────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "Resultado do Assessment Psicológico"
    ws["A1"].font = title_font
    ws["A3"] = "Candidato"
    ws["B3"] = candidate.full_name if candidate else ""
    ws["A4"] = "E-mail"
    ws["B4"] = candidate.email if candidate else ""
    ws["A5"] = "Concluído em"
    ws["B5"] = response.completed_at.strftime("%d/%m/%Y %H:%M") if response.completed_at else ""
    for r in (3, 4, 5):
        ws.cell(row=r, column=1).font = Font(bold=True)
    autosize(ws, [22, 50])

    scores = response.scores or {}

    # Scores summary block
    row = 8
    ws.cell(row=row, column=1, value="Scores consolidados").font = Font(bold=True, size=12)
    row += 2

    if "big_five" in scores:
        ws.cell(row=row, column=1, value="Big Five").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Fator")
        ws.cell(row=row, column=2, value="Score (%)")
        ws.cell(row=row, column=3, value="Bruto / Máx")
        write_headers(ws, ["Fator", "Score (%)", "Bruto / Máx"])
        row += 1
        for factor, data in scores["big_five"].items():
            ws.cell(row=row, column=1, value=factor)
            ws.cell(row=row, column=2, value=f"{round(data.get('score', 0) * 100, 1)}%")
            ws.cell(row=row, column=3, value=f"{data.get('raw', '')}/{data.get('max', '')}")
            row += 1
        row += 1

    for label, key in [("Stress", "stress"), ("Depressão", "depression"), ("Quociente de Adversidade", "qa")]:
        if key in scores:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value=f"{scores[key].get('score', '')}/{scores[key].get('max', '')}")
            ws.cell(row=row, column=2, value=scores[key].get("level", ""))
            row += 2

    if "motivators" in scores:
        ws.cell(row=row, column=1, value="Motivadores").font = Font(bold=True)
        row += 1
        for group, m in scores["motivators"].items():
            ws.cell(row=row, column=1, value=group)
            ws.cell(row=row, column=2, value=f"{m.get('score', '')}/{m.get('max', '')}")
            ws.cell(row=row, column=3, value=f"{round(m.get('pct', 0) * 100, 1)}%")
            row += 1
        row += 1

    if "disc" in scores:
        ws.cell(row=row, column=1, value="DISC").font = Font(bold=True)
        row += 1
        for dim, d in scores["disc"].items():
            ws.cell(row=row, column=1, value=f"{dim} - {d.get('name', '')}")
            ws.cell(row=row, column=2, value=f"{d.get('score', '')}/{d.get('max', '')}")
            ws.cell(row=row, column=3, value=d.get("level", ""))
            row += 1

    # ── Big Five ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("BigFive")
    ws["A1"] = "Big Five — Traços de Personalidade"
    ws["A1"].font = title_font
    ws["A2"] = "Escala: 1 (Não me descreve) → 4 (Descreve-me na média) → 7 (Descreve-me muito bem)"
    ws["A2"].font = sub_font
    ws.append([])
    ws.append(["Nº", "Pergunta", "Resposta", "Fator", "Sub-fator"])
    write_headers(ws, ["Nº", "Pergunta", "Resposta", "Fator", "Sub-fator"])
    for q in BIG_FIVE_QUESTIONS:
        ans = response.big_five_answers.get(str(q["num"]), "")
        ws.append([q["num"], q["text"], ans, q["factor"], q["sub_factor"]])
    autosize(ws, [6, 70, 10, 22, 30])
    for r in ws.iter_rows(min_row=5):
        r[1].alignment = wrap

    # ── Stress ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Stress")
    ws["A1"] = "Nível de Stress"
    ws["A1"].font = title_font
    ws["A2"] = "Escala: 0 Nunca · 1 Quase nunca · 2 Algumas vezes · 3 Com bastante frequência · 4 Quase sempre"
    ws["A2"].font = sub_font
    ws.append([])
    ws.append(["Nº", "Pergunta", "Resposta"])
    write_headers(ws, ["Nº", "Pergunta", "Resposta"])
    stress_lbl = {0: "Nunca", 1: "Quase nunca", 2: "Algumas vezes", 3: "Com bastante frequência", 4: "Quase sempre"}
    for q in STRESS_QUESTIONS:
        v = response.stress_answers.get(str(q["num"]), "")
        try:
            ans = f"{v} ({stress_lbl[int(v)]})"
        except Exception:
            ans = str(v)
        ws.append([q["num"], q["text"], ans])
    autosize(ws, [6, 70, 28])
    for r in ws.iter_rows(min_row=5):
        r[1].alignment = wrap

    # ── Depression ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Depressao")
    ws["A1"] = "Sintomático de Depressão"
    ws["A1"].font = title_font
    ws["A2"] = "Cada questão tem 4 opções (índice 0-3). Mostrado o índice + texto da opção escolhida."
    ws["A2"].font = sub_font
    ws.append([])
    ws.append(["Chave", "Pergunta", "Resposta", "Opções disponíveis"])
    write_headers(ws, ["Chave", "Pergunta", "Resposta", "Opções disponíveis"])
    for q in DEPRESSION_QUESTIONS:
        v = response.depression_answers.get(q["key"], "")
        try:
            idx = int(v)
            ans = f"{idx} — {q['options'][idx]}"
        except Exception:
            ans = str(v)
        opts = " | ".join(f"{i}={o}" for i, o in enumerate(q["options"]))
        ws.append([q["key"], q["text"], ans, opts])
    autosize(ws, [8, 60, 50, 80])
    for r in ws.iter_rows(min_row=5):
        for c in (1, 2, 3):
            r[c].alignment = wrap

    # ── Motivators ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Motivadores")
    ws["A1"] = "Motivadores"
    ws["A1"].font = title_font
    ws["A2"] = "Escala: 0 Nunca · 1 Quase nunca · 2 Poucas vezes · 3 Normalmente · 4 Quase sempre · 5 O tempo todo"
    ws["A2"].font = sub_font
    ws.append([])
    ws.append(["Grupo", "Nº", "Pergunta", "Resposta"])
    write_headers(ws, ["Grupo", "Nº", "Pergunta", "Resposta"])
    mot_lbl = {0: "Nunca", 1: "Quase nunca", 2: "Poucas vezes", 3: "Normalmente", 4: "Quase sempre", 5: "O tempo todo"}
    for group, questions in MOTIVATORS_GROUPS.items():
        for q in questions:
            v = response.motivators_answers.get(str(q["num"]), "")
            try:
                ans = f"{v} ({mot_lbl[int(v)]})"
            except Exception:
                ans = str(v)
            ws.append([group, q["num"], q["text"], ans])
    autosize(ws, [22, 6, 70, 28])
    for r in ws.iter_rows(min_row=5):
        r[2].alignment = wrap

    # ── DISC ─────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("DISC")
    ws["A1"] = "DISC"
    ws["A1"].font = title_font
    ws["A2"] = "Para cada par de adjetivos, o candidato escolheu um valor (0-5). 0 = totalmente o da esquerda; 5 = totalmente o da direita."
    ws["A2"].font = sub_font
    ws.append([])
    ws.append(["Dimensão", "Item", "Esquerda", "Direita", "Resposta"])
    write_headers(ws, ["Dimensão", "Item", "Esquerda", "Direita", "Resposta"])
    for dim, info in DISC_DIMENSIONS.items():
        values = response.disc_answers.get(dim, [])
        for idx, pair in enumerate(info["pairs"]):
            v = values[idx] if idx < len(values) else ""
            ws.append([f"{dim} - {info['name']}", idx + 1, pair[0], pair[1], v])
    autosize(ws, [22, 8, 30, 30, 12])

    # ── QA ───────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("QA")
    ws["A1"] = "Quociente de Adversidade"
    ws["A1"].font = title_font
    ws["A2"] = "Escala: 1 Nunca · 2 Pouco · 3 Médio · 4 Na maioria das vezes · 5 Sempre"
    ws["A2"].font = sub_font
    ws.append([])
    ws.append(["Nº", "Pergunta", "Resposta"])
    write_headers(ws, ["Nº", "Pergunta", "Resposta"])
    qa_lbl = {1: "Nunca", 2: "Pouco", 3: "Médio", 4: "Na maioria das vezes", 5: "Sempre"}
    for q in QA_QUESTIONS:
        v = response.qa_answers.get(str(q["num"]), "")
        try:
            ans = f"{v} ({qa_lbl[int(v)]})"
        except Exception:
            ans = str(v)
        ws.append([q["num"], q["text"], ans])
    autosize(ws, [6, 70, 28])
    for r in ws.iter_rows(min_row=5):
        r[1].alignment = wrap

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    cand_name = (candidate.full_name if candidate else str(app_id)).replace(" ", "_")
    filename = f"psych-results-{cand_name}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
